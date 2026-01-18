#!/usr/bin/env python3
"""
Excel Formatter - Format Excel output with specific column requirements
"""

import pandas as pd
import yaml
import os
import re
import logging
from datetime import datetime, timedelta
from typing import List, Dict
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openai import OpenAI

logger = logging.getLogger(__name__)


class ExcelFormatter:
    """Format and write Excel files with custom columns"""
    
    def __init__(self, config_path: str = "config.yml"):
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = yaml.safe_load(f)
        
        self.output_dir = self.config['excel']['output_dir']
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Initialize AI client for deduplication (works for both cloud and local AI)
        deepseek_config = self.config.get('deepseek', {})
        if deepseek_config.get('api_key') or deepseek_config.get('api_base'):
            self.ai_client = OpenAI(
                api_key=deepseek_config.get('api_key', 'local-key'),
                base_url=deepseek_config.get('api_base', 'https://api.deepseek.com')
            )
            # Support both 'model' and 'chat_model' config keys for compatibility
            self.ai_model = deepseek_config.get('chat_model', deepseek_config.get('model', 'deepseek-chat'))
        else:
            self.ai_client = None
    
    def get_next_monday_filename(self, end_date: datetime) -> str:
        """
        Get filename based on next Monday after period end + current time
        Format: YYMMDD_HHMMSS (e.g., 251215_103045 for 15 Dec 2025, 10:30:45)
        """
        # Find next Monday
        days_until_monday = (7 - end_date.weekday()) % 7
        if days_until_monday == 0:
            days_until_monday = 7
        next_monday = end_date + timedelta(days=days_until_monday)
        
        # Format as YYMMDD_HHMMSS (date + current time)
        current_time = datetime.now().strftime('%H%M%S')
        return f"{next_monday.strftime('%y%m%d')}_{current_time}"
    
    def extract_source(self, article: Dict) -> str:
        """Extract source from article content"""
        # First check if source was extracted from article page
        source = article.get('source', 'Company C')
        
        # Don't expose internal source names - if we have a real source name, use it
        if source and source not in ['Company A', 'Company B', 'Company C', '852.house']:
            return source  # Keep actual news source names (e.g., 經濟日報, 星島日報)
        
        # If source is still Company C, check if it's in the known sources list
        if source == 'Company C':
            # Check tags as fallback
            tags = article.get('tags', [])
            sources = self.config.get('sources', [])
            
            for tag in tags:
                for known_source in sources:
                    if known_source in tag:
                        return known_source
            
            # Default to 852.house only if we really can't find anything
            return "852.house"
        
        # If source is 852.house or something else, return as is
        return source if source else "852.house"
    
    def deduplicate_transactions(self, articles: List[Dict]) -> List[Dict]:
        """
        Deduplicate transactions based on property + date
        Keep the one with most complete information
        Add dedup_flag for manual review
        """
        from collections import defaultdict
        
        # Group by property name + date
        groups = defaultdict(list)
        for article in articles:
            details = article.get('details', {})
            # Create key from property name + date
            property_name = details.get('property', '').strip()
            date = details.get('date', '')
            # Normalize property name (remove spaces, convert to lower)
            key = f"{property_name.replace(' ', '').lower()}_{date}"
            groups[key].append(article)
        
        # For each group, keep the most complete one
        deduped = []
        for key, group in groups.items():
            if len(group) == 1:
                # No duplicates
                group[0]['dedup_flag'] = ''
                deduped.append(group[0])
            else:
                # Multiple articles for same property+date
                # Score by completeness
                def completeness_score(article):
                    details = article.get('details', {})
                    score = 0
                    for field in ['district', 'floor', 'unit', 'price', 'area', 'unit_price', 
                                  'buyer', 'seller', 'yield_rate']:
                        if details.get(field) and details.get(field) != 'N/A':
                            score += 1
                    return score
                
                # Sort by completeness (highest first)
                sorted_group = sorted(group, key=completeness_score, reverse=True)
                
                # Keep the most complete one, mark for review
                best = sorted_group[0]
                best['dedup_flag'] = f'REVIEW: {len(group)} duplicates found'
                deduped.append(best)
        
        return deduped
    
    def format_transactions(self, articles: List[Dict], filename: str) -> pd.DataFrame:
        """Format transactions sheet"""
        # Deduplicate first
        deduped_articles = self.deduplicate_transactions(articles)
        
        # Filter out rows with N/A property names AND missing area
        valid_articles = []
        for a in deduped_articles:
            details = a.get('details', {})
            property_name = details.get('property', 'N/A')
            area = details.get('area', 'N/A')
            
            # Must have property name AND area
            if property_name != 'N/A' and area != 'N/A':
                # Ensure area is a valid number
                try:
                    float(str(area).replace(',', ''))
                    valid_articles.append(a)
                except (ValueError, AttributeError):
                    # Area is not a valid number, skip
                    continue
        
        data = []
        for idx, article in enumerate(valid_articles, 1):
            details = article.get('details', {})
            
            # Convert numeric fields to proper format
            def to_numeric(value, default='N/A'):
                if value == 'N/A' or value is None:
                    return default
                try:
                    # Remove commas and convert
                    num_str = str(value).replace(',', '').strip()
                    if num_str and num_str != 'N/A':
                        return float(num_str)
                    return default
                except (ValueError, AttributeError):
                    return default
            
            price = to_numeric(details.get('price', 'N/A'), 'N/A')
            area = to_numeric(details.get('area', 'N/A'), 'N/A')
            unit_price = to_numeric(details.get('unit_price', 'N/A'), 'N/A')
            yield_rate = details.get('yield_rate', 'N/A')
            if yield_rate != 'N/A' and yield_rate is not None:
                try:
                    yield_rate = float(yield_rate)
                except (ValueError, TypeError):
                    yield_rate = 'N/A'
            
            # Determine area_basis based on asset_type
            asset_type = details.get('asset_type', 'N/A')
            if asset_type in ['住宅', '洋房']:
                area_basis = 'NFA'
            elif asset_type in ['寫字樓', '商鋪', '商舖', '工廈', '酒店', '停車位']:
                area_basis = 'GFA'
            else:
                area_basis = 'NFA'  # Default to NFA
            
            row = {
                'No.': idx,
                'Date': details.get('date', 'N/A'),
                'District': details.get('district', 'N/A'),
                'Property': details.get('property', article.get('title', '')[:50]),
                '': '',  # Empty column after Property
                'Asset type': details.get('asset_type', 'N/A'),
                'Floor': details.get('floor', 'N/A'),
                'Unit': details.get('unit', 'N/A'),
                'Nature': details.get('nature', 'N/A'),
                'Transaction price': price,
                'Area basis': area_basis,
                'Unit basis': 'sqft',
                'Area/unit': area,
                'Unit price': unit_price,
                'Yield': yield_rate,
                'Seller/Landlord': details.get('seller', 'N/A'),
                'Buyer/Tenant': details.get('buyer', 'N/A'),
                'Source': self.extract_source(article),
                'URL': article.get('url', ''),
                'Filename': filename,
                'Dedup Flag': article.get('dedup_flag', '')
            }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def format_centaline(self, transactions: List[Dict], filename: str) -> pd.DataFrame:
        """Format Centaline/Midland transactions sheet"""
        data = []
        
        def to_numeric(value, default='N/A'):
            """Convert value to numeric, return default if not valid"""
            if value == 'N/A' or value is None:
                return default
            try:
                num_str = str(value).replace(',', '').strip()
                if num_str and num_str != 'N/A':
                    return float(num_str)
                return default
            except (ValueError, AttributeError):
                return default
        
        # First pass: collect valid rows (property name must not be empty)
        valid_rows = []
        
        for trans in transactions:
            property_name = trans.get('property', '').strip()
            
            # Skip if property name is empty or N/A
            if not property_name or property_name == 'N/A':
                continue
            
            # Determine source and category
            source = trans.get('source', 'Company A')
            category = trans.get('category', 'Residential')
            
            # Identify source clearly
            if 'Midland' in source or 'Company B' in source or source == 'Midland':
                source_name = 'Midland'
                if not category or category == 'Residential':
                    category = 'Commercial'
            elif 'Centaline' in source or 'Company A' in source or '中原' in source:
                source_name = 'Centaline'
                if not category or category == 'Commercial':
                    category = 'Residential'
            else:
                source_name = source
            
            # Convert to numeric values
            area = to_numeric(trans.get('area', trans.get('area_unit', 'N/A')), 'N/A')
            price = to_numeric(trans.get('price_numeric', trans.get('price', 'N/A')), 'N/A')
            unit_price = to_numeric(trans.get('unit_price', 'N/A'), 'N/A')
            
            # Determine area_basis based on asset_type
            asset_type = trans.get('asset_type', '住宅')
            if asset_type in ['住宅', '洋房']:
                area_basis = 'NFA'
            elif asset_type in ['寫字樓', '商鋪', '商舖', '工廈', '工商', '酒店', '停車位']:
                area_basis = 'GFA'
            else:
                area_basis = 'NFA'
            
            # Normalize nature to English
            nature = trans.get('nature', 'Sales')
            if nature in ['租', 'L', 'Lease', 'LEASE']:
                nature = 'Lease'
            elif nature in ['售', 'S', 'Sales', 'SALES', 'Sale']:
                nature = 'Sales'
            else:
                nature = 'Sales'  # Default
            
            # Store transaction data (without No. yet)
            valid_rows.append({
                'Date': trans.get('date', 'N/A'),
                'District': trans.get('district', 'N/A'),
                'Asset type': asset_type,
                'Property': property_name,
                'Floor': trans.get('floor', 'N/A'),
                'Unit': trans.get('unit', 'N/A'),
                'Area basis': area_basis,
                'Unit basis': 'sqft',
                'Area/Unit': area,
                'Transaction Price': price,
                'Unit Price': unit_price,
                'Nature': nature,
                'Category': category,
                'Source': source_name,
                'Filename': filename
            })
        
        # Second pass: add row numbers and build final data with correct column order
        for idx, trans_data in enumerate(valid_rows, 1):
            row = {
                'No.': idx,
                'Date': trans_data['Date'],
                'District': trans_data['District'],
                'Asset type': trans_data['Asset type'],
                'Property': trans_data['Property'],
                'Floor': trans_data['Floor'],
                'Unit': trans_data['Unit'],
                'Area basis': trans_data['Area basis'],
                'Unit basis': trans_data['Unit basis'],
                'Area/Unit': trans_data['Area/Unit'],
                'Transaction Price': trans_data['Transaction Price'],
                'Unit Price': trans_data['Unit Price'],
                'Nature': trans_data['Nature'],
                'Category': trans_data['Category'],
                'Source': trans_data['Source'],
                'Filename': trans_data['Filename']
            }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def format_news(self, articles: List[Dict], filename: str) -> pd.DataFrame:
        """Format news sheet - renumber after all filtering"""
        data = []
        
        # Renumber starting from 1 after all filters applied
        for idx, article in enumerate(articles, 1):
            details = article.get('details', {})
            
            row = {
                'No.': idx,
                'Date': details.get('date', article.get('date', 'N/A')),
                'Source': self.extract_source(article),
                'Asset type': details.get('asset_category', 'General'),
                'Topic': details.get('topic', article.get('title', '')),
                'Summary': details.get('summary', 'N/A'),
                'URL': article.get('url', ''),
                'Filename': filename
            }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def format_worksheet(self, worksheet, is_transaction: bool = True):
        """Apply formatting to worksheet"""
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Column widths for transactions
        if is_transaction:
            widths = {
                'A': 6,   # No.
                'B': 12,  # Date
                'C': 12,  # District
                'D': 30,  # Property
                'E': 3,   # Empty column
                'F': 10,  # Asset type
                'G': 8,   # Floor
                'H': 8,   # Unit
                'I': 8,   # Nature
                'J': 15,  # Transaction price
                'K': 10,  # Area basis
                'L': 10,  # Unit basis
                'M': 10,  # Area/unit
                'N': 12,  # Unit price
                'O': 10,  # Yield
                'P': 20,  # Seller/Landlord
                'Q': 20,  # Buyer/Tenant
                'R': 12,  # Source
                'S': 15,  # URL
                'T': 10,  # Filename
                'U': 25   # Dedup Flag
            }
        else:
            # News columns
            widths = {
                'A': 6,   # No.
                'B': 12,  # Date
                'C': 12,  # Source
                'D': 12,  # Asset type
                'E': 40,  # Topic
                'F': 60,  # Summary
                'G': 15,  # URL
                'H': 10   # Filename
            }
        
        for col, width in widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Text wrapping
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Freeze header
        worksheet.freeze_panes = 'A2'
    
    def _format_centaline_sheet(self, worksheet):
        """Format Centaline worksheet"""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Column widths
        widths = {
            'A': 6,   # No.
            'B': 12,  # Date
            'C': 12,  # District
            'D': 10,  # Asset type
            'E': 30,  # Property
            'F': 8,   # Floor
            'G': 8,   # Unit
            'H': 10,  # Area basis
            'I': 10,  # Unit basis
            'J': 10,  # Area/Unit
            'K': 15,  # Transaction Price
            'L': 12,  # Unit Price
            'M': 8,   # Nature
            'N': 12,  # Category
            'O': 12,  # Source
            'P': 10   # Filename
        }
        
        for col, width in widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Text wrapping
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        worksheet.freeze_panes = 'A2'
    
    def ai_deduplicate_news(self, articles: List[Dict]) -> List[Dict]:
        """
        Use AI to deduplicate highly similar news articles by comparing topic+summary pairs
        More aggressive deduplication - compares ALL articles, not just within date groups
        
        Args:
            articles: List of news articles with details
            
        Returns:
            List of deduplicated articles
        """
        if not self.ai_client or len(articles) <= 1:
            return articles
        
        # Compare ALL articles for better deduplication
        unique_articles = []
        total_compared = 0
        total_removed = 0
        
        for article1 in articles:
            topic1 = article1.get('details', {}).get('topic', '')
            summary1 = article1.get('details', {}).get('summary', '')
            
            if not topic1:
                unique_articles.append(article1)
                continue
            
            is_duplicate = False
            for article2 in unique_articles:
                topic2 = article2.get('details', {}).get('topic', '')
                summary2 = article2.get('details', {}).get('summary', '')
                
                if not topic2:
                    continue
                
                # Quick pre-check: if topics are very different (< 30% overlap), skip AI
                topic1_words = set(topic1.replace(' ', '').replace('\u3000', '').lower())
                topic2_words = set(topic2.replace(' ', '').replace('\u3000', '').lower())
                if topic1_words and topic2_words:
                    # Calculate character overlap
                    overlap = len(topic1_words & topic2_words) / max(len(topic1_words), len(topic2_words))
                    if overlap < 0.3:
                        continue  # Skip AI call for very different topics
                
                # Use AI to check similarity for potentially similar articles
                total_compared += 1
                if self._are_articles_similar(topic1, summary1, topic2, summary2):
                    is_duplicate = True
                    total_removed += 1
                    break
            
            if not is_duplicate:
                unique_articles.append(article1)
        
        if total_compared > 0:
            print(f"    → AI compared {total_compared} pairs, removed {total_removed} duplicates")
        
        return unique_articles
    
    def rank_and_filter_news(self, articles: List[Dict], target_count: int = 20) -> List[Dict]:
        """
        Rank news articles by market relevance and keep only top articles
        
        Args:
            articles: List of news articles
            target_count: Target number of articles to keep (default 20)
            
        Returns:
            List of top-ranked articles
        """
        if not self.ai_client or len(articles) <= target_count:
            return articles
        
        print(f"    → Scoring {len(articles)} articles for market relevance...")
        
        # Score each article
        scored_articles = []
        for article in articles:
            topic = article.get('details', {}).get('topic', '')
            summary = article.get('details', {}).get('summary', '')
            
            if not topic:
                continue
            
            # Get relevance score from AI
            score = self._score_market_relevance(topic, summary)
            scored_articles.append((score, article))
        
        # Sort by score (highest first) and take top articles
        scored_articles.sort(key=lambda x: x[0], reverse=True)
        
        # Keep top target_count articles (or minimum 15)
        keep_count = max(min(target_count, len(scored_articles)), 15)
        top_articles = [article for score, article in scored_articles[:keep_count] if score >= 6]
        
        # If we filtered too aggressively and have less than 15, keep more
        if len(top_articles) < 15 and len(scored_articles) >= 15:
            top_articles = [article for score, article in scored_articles[:15]]
        
        return top_articles
    
    def _score_market_relevance(self, topic: str, summary: str) -> int:
        """
        Score article's relevance to HK market valuation (0-10)
        
        Args:
            topic: Article topic/title
            summary: Article summary
            
        Returns:
            Score from 0-10 (10 = most relevant to market valuation)
        """
        prompt = f"""請評分以下香港地產新聞對整體市場估值的重要性和相關性。

評分標準 (0-10分):
10分: 重大政策變動、利率調整、整體市場數據/趨勢，對市場估值有直接重大影響
8-9分: 重要市場數據、土地供應、大型發展商動向，有明確市場影響
6-7分: 一般市場新聞、區域數據、次要政策，有一定參考價值
4-5分: 個別項目新聞、地區性消息，市場影響有限
2-3分: 評論文章、個別案例、零散資訊，參考價值低
0-1分: 與市場估值無關、質素問題、個人故事、社區瑣事

標題: {topic}
摘要: {summary}

請只回答一個數字(0-10)，不要其他說明。"""

        try:
            response = self.ai_client.chat.completions.create(
                model=self.ai_model,
                messages=[
                    {"role": "system", "content": "你是香港地產市場分析專家，專門評估新聞對市場估值的重要性。"},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,
                max_tokens=10
            )
            
            score_text = response.choices[0].message.content.strip()
            # Extract number from response
            import re
            score_match = re.search(r'\d+', score_text)
            if score_match:
                score = int(score_match.group())
                return min(10, max(0, score))  # Clamp to 0-10
            else:
                return 5  # Default moderate score
                
        except Exception as e:
            logger.error(f"Error scoring article: {e}")
            return 5  # Default moderate score
    
    def _are_articles_similar(self, topic1: str, summary1: str, topic2: str, summary2: str) -> bool:
        """
        Use AI to determine if two articles are highly similar
        
        Args:
            topic1, summary1: First article's topic and summary
            topic2, summary2: Second article's topic and summary
            
        Returns:
            True if articles are highly similar (duplicates), False otherwise
        """
        if not self.ai_client:
            return False
        
        prompt = f"""請判斷以下兩則香港地產新聞是否高度相似（內容基本相同，只是表述略有不同）。

新聞1:
標題: {topic1}
摘要: {summary1}

新聞2:
標題: {topic2}
摘要: {summary2}

請判斷這兩則新聞是否講述相同的事件或內容。如果它們高度相似（例如：同一事件的不同報道、同一數據的不同表述），請回答"是"。如果它們是不同的新聞內容，請回答"否"。

只回答"是"或"否"，不要添加其他說明。"""

        try:
            response = self.ai_client.chat.completions.create(
                model=self.ai_model,
                messages=[
                    {"role": "system", "content": "你是一個新聞去重專家。請準確判斷兩則新聞是否高度相似。"},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=10
            )
            
            result = response.choices[0].message.content.strip().lower()
            return '是' in result or 'yes' in result or result == 'y'
            
        except Exception as e:
            logger.debug(f"Error in AI deduplication: {e}")
            return False
    
    def format_new_property(self, articles: List[Dict], filename: str) -> pd.DataFrame:
        """Format new property sheet"""
        data = []
        
        def to_numeric(value, default='N/A'):
            """Convert value to numeric, return default if not valid"""
            if value == 'N/A' or value is None:
                return default
            try:
                num_str = str(value).replace(',', '').strip()
                if num_str and num_str != 'N/A':
                    return float(num_str)
                return default
            except (ValueError, AttributeError):
                return default
        
        for idx, article in enumerate(articles, 1):
            details = article.get('details', {})
            title = article.get('title', '')
            content = article.get('full_content', article.get('description', ''))
            
            # Convert numeric fields
            price_min = to_numeric(details.get('price_min', details.get('price', 'N/A')), 'N/A')
            price_max = to_numeric(details.get('price_max', details.get('price', 'N/A')), 'N/A')
            area_min = to_numeric(details.get('area_min', details.get('area', 'N/A')), 'N/A')
            area_max = to_numeric(details.get('area_max', details.get('area', 'N/A')), 'N/A')
            unit_price_min = to_numeric(details.get('unit_price_min', details.get('unit_price', 'N/A')), 'N/A')
            unit_price_max = to_numeric(details.get('unit_price_max', details.get('unit_price', 'N/A')), 'N/A')
            unit_price_avg = to_numeric(details.get('unit_price_avg', details.get('unit_price', 'N/A')), 'N/A')
            
            row = {
                'No.': idx,
                'Date': details.get('date', article.get('date', 'N/A')),
                'District': details.get('district', 'N/A'),
                'Property': details.get('property', title[:50]),
                'Asset type': details.get('asset_type', 'N/A'),
                'Floor': details.get('floor', 'N/A'),
                'Unit': details.get('unit', 'N/A'),
                'Nature': details.get('nature', 'Sales'),
                'Transaction Price_Min': price_min,
                'Transaction Price_Max': price_max,
                'Area basis': details.get('area_basis', 'NFA'),
                'Unit basis': 'sqft',
                'Area_Min': area_min,
                'Area_Max': area_max,
                'Unit Price_Min': unit_price_min,
                'Unit Price_Max': unit_price_max,
                'Unit Price_Avg': unit_price_avg,
                'Seller/Landlord': details.get('seller', 'N/A'),
                'Source': self.extract_source(article),
                'URL': article.get('url', ''),
                'Filename': filename,
                'Content': content[:500] if content else 'N/A'  # First 500 chars
            }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def write_excel(self, transactions: List[Dict], news: List[Dict], 
                   centaline: List[Dict], midland: List[Dict], 
                   start_date: datetime, end_date: datetime) -> str:
        """Write formatted Excel file"""
        # Get filename with timestamp for Excel file
        excel_filename = self.get_next_monday_filename(end_date)
        filepath = os.path.join(self.output_dir, f"property_report_{excel_filename}.xlsx")
        
        # Filename for tabs (just date, no time)
        days_until_monday = (7 - end_date.weekday()) % 7
        if days_until_monday == 0:
            days_until_monday = 7
        next_monday = end_date + timedelta(days=days_until_monday)
        tab_filename = next_monday.strftime('%y%m%d')
        
        print(f"\n  → Creating Excel file: {filepath}")
        print(f"  → Tab filename code: {tab_filename}")
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Transactions sheet (always create)
            df_trans = self.format_transactions(transactions, tab_filename) if transactions else pd.DataFrame()
            if not df_trans.empty:
                df_trans.to_excel(writer, sheet_name='major_trans', index=False)
                self.format_worksheet(writer.book['major_trans'], is_transaction=True)
                print(f"  → major_trans: {len(df_trans)} rows")
            else:
                df_trans = pd.DataFrame(columns=['No.', 'Date', 'District', 'Property', '', 'Asset type', 
                                                 'Floor', 'Unit', 'Nature', 'Transaction price', 
                                                 'Area basis', 'Unit basis', 'Area/unit', 'Unit price', 'Yield', 
                                                 'Seller/Landlord', 'Buyer/Tenant', 'Source', 'URL', 
                                                 'Filename', 'Dedup Flag'])
                df_trans.to_excel(writer, sheet_name='major_trans', index=False)
                self.format_worksheet(writer.book['major_trans'], is_transaction=True)
                print(f"  → major_trans: 0 rows (empty)")
            
            # News sheet - deduplicate by topic (normalize by removing spaces)
            if news:
                original_count = len(news)
                seen_topics_normalized = set()
                unique_news = []
                for article in news:
                    topic = article.get('details', {}).get('topic', '')
                    if topic:
                        # Normalize topic by removing all spaces for comparison
                        topic_normalized = topic.replace(' ', '').replace('\u3000', '')  # Remove regular spaces and full-width spaces
                        if topic_normalized and topic_normalized not in seen_topics_normalized:
                            seen_topics_normalized.add(topic_normalized)
                            unique_news.append(article)
                news = unique_news
                print(f"  → Deduplicated news (space-normalized): {len(unique_news)} unique (removed {original_count - len(unique_news)} duplicates)")
                
                # AI-based deduplication for highly similar articles BEFORE filtering
                if len(news) > 1 and self.ai_client:
                    print(f"  → AI deduplication: checking for highly similar articles...")
                    news_before_ai = len(news)
                    news = self.ai_deduplicate_news(news)
                    print(f"  → After AI deduplication: {len(news)} unique articles (removed {news_before_ai - len(news)} similar)")
                
                # Rank and filter to top 15-20 most market-relevant articles
                if len(news) > 20:
                    print(f"  → Too many news ({len(news)}). Ranking by market relevance (target: 15-20)...")
                    news = self.rank_and_filter_news(news, target_count=20)
                    print(f"  → Kept top {len(news)} most market-relevant articles")
                elif len(news) > 15:
                    # Even if between 15-20, still rank to ensure quality
                    print(f"  → Ranking {len(news)} news by market relevance for quality...")
                    news = self.rank_and_filter_news(news, target_count=min(20, len(news)))
                    print(f"  → Kept top {len(news)} most market-relevant articles")
            
            df_news = self.format_news(news, tab_filename) if news else pd.DataFrame()
            if not df_news.empty:
                df_news.to_excel(writer, sheet_name='news', index=False)
                self.format_worksheet(writer.book['news'], is_transaction=False)
                print(f"  → news: {len(df_news)} rows")
            else:
                df_news = pd.DataFrame(columns=['No.', 'Date', 'Source', 'Asset type', 'Topic', 'Summary', 'URL', 'Filename'])
                df_news.to_excel(writer, sheet_name='news', index=False)
                self.format_worksheet(writer.book['news'], is_transaction=False)
                print(f"  → news: 0 rows (empty)")
            
            # Trans_Commercial sheet - combine Centaline + Midland
            all_commercial = []
            if centaline:
                all_commercial.extend(centaline)
            if midland:
                all_commercial.extend(midland)
            
            if all_commercial:
                df_commercial = self.format_centaline(all_commercial, tab_filename)
                df_commercial.to_excel(writer, sheet_name='Trans_Commercial', index=False)
                self._format_centaline_sheet(writer.book['Trans_Commercial'])
                print(f"  → Trans_Commercial: {len(df_commercial)} rows")
            else:
                df_commercial = pd.DataFrame(columns=['No.', 'Date', 'District', 'Asset type', 'Property', 
                                                     'Floor', 'Unit', 'Area basis', 'Unit basis', 'Area/Unit', 
                                                     'Transaction Price', 'Unit Price', 'Nature', 'Category', 
                                                     'Source', 'Filename'])
                df_commercial.to_excel(writer, sheet_name='Trans_Commercial', index=False)
                self._format_centaline_sheet(writer.book['Trans_Commercial'])
                print(f"  → Trans_Commercial: 0 rows (empty)")
            
            # New Property sheet (empty template - not used)
            df_new_prop = pd.DataFrame(columns=['No.', 'Date', 'District', 'Property', 'Asset type', 
                                               'Floor', 'Unit', 'Nature', 'Transaction Price_Min', 
                                               'Transaction Price_Max', 'Area basis', 'Unit basis', 
                                               'Area_Min', 'Area_Max', 'Unit Price_Min', 'Unit Price_Max', 
                                               'Unit Price_Avg', 'Seller/Landlord', 'Source', 'URL', 
                                               'Filename', 'Content'])
            df_new_prop.to_excel(writer, sheet_name='new_property', index=False)
            self.format_worksheet(writer.book['new_property'], is_transaction=True)
            print(f"  → new_property: 0 rows (template)")
        
        print(f"\n✅ Excel file created: {filepath}")
        return filepath

