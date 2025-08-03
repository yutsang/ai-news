import pandas as pd
import os
from datetime import datetime
from typing import Dict, List
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from config import REPORT_CONFIG, BIG_DEALS_BASELINE

logger = logging.getLogger(__name__)

class ExcelReportGenerator:
    def __init__(self):
        self.output_dir = REPORT_CONFIG['output_dir']
        self.date_format = REPORT_CONFIG['date_format']
        self.filename_template = REPORT_CONFIG['excel_filename_template']
        
        # Create output directory if it doesn't exist
        os.makedirs(self.output_dir, exist_ok=True)
    
    def _auto_adjust_column_widths(self, ws):
        """Auto-adjust column widths safely."""
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            # Find the first cell with a column_letter attribute
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            if column_letter:
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_excel_report(self, report_data: Dict, transactions: List[Dict], news: List[Dict]) -> str:
        """
        Create Excel report with the specified format.
        """
        # Get current date for filename
        current_date = datetime.now().strftime(self.date_format)
        filename = self.filename_template.format(date=current_date)
        filepath = os.path.join(self.output_dir, filename)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_executive_summary_sheet(wb, report_data)
        self._create_transactions_sheet(wb, transactions)
        self._create_news_sheet(wb, news)
        self._create_statistics_sheet(wb, report_data)
        self._create_big_deals_sheet(wb, transactions)
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"Excel report saved: {filepath}")
        
        return filepath
    
    def _create_executive_summary_sheet(self, wb: Workbook, report_data: Dict):
        """Create executive summary sheet."""
        ws = wb.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "VMTA Market Updates - Executive Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Period
        ws['A3'] = f"Period: {report_data['period']['start_date']} to {report_data['period']['end_date']}"
        ws['A3'].font = Font(bold=True)
        
        # Executive Summary
        ws['A5'] = "Executive Summary"
        ws['A5'].font = Font(size=14, bold=True)
        ws['A6'] = report_data['executive_summary']['executive_summary']
        ws['A6'].alignment = Alignment(wrap_text=True)
        
        # Key Highlights
        ws['A8'] = "Key Highlights"
        ws['A8'].font = Font(size=14, bold=True)
        for i, highlight in enumerate(report_data['executive_summary']['key_highlights'], 1):
            ws[f'A{8+i}'] = f"• {highlight}"
        
        # Market Outlook
        start_row = 8 + len(report_data['executive_summary']['key_highlights']) + 2
        ws[f'A{start_row}'] = "Market Outlook"
        ws[f'A{start_row}'].font = Font(size=14, bold=True)
        ws[f'A{start_row+1}'] = report_data['executive_summary']['market_outlook']
        ws[f'A{start_row+1}'].alignment = Alignment(wrap_text=True)
        
        # Recommendations
        start_row += 3
        ws[f'A{start_row}'] = "Recommendations"
        ws[f'A{start_row}'].font = Font(size=14, bold=True)
        for i, rec in enumerate(report_data['executive_summary']['recommendations'], 1):
            ws[f'A{start_row+i}'] = f"• {rec}"
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws)
    
    def _create_transactions_sheet(self, wb: Workbook, transactions: List[Dict]):
        """Create transactions sheet."""
        ws = wb.create_sheet("Market Transactions")
        
        # Headers - start from row 1
        headers = [
            'Date', 'Source', 'Property Name', 'Location', 'Asset Type',
            'Transaction Type', 'Value (HKD)', 'Big Deal', 'URL'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data - start from row 2
        for row, transaction in enumerate(transactions, 2):
            transaction_data = transaction.get('transaction_data', {})
            
            ws.cell(row=row, column=1, value=transaction['date'].strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=transaction['source'])
            ws.cell(row=row, column=3, value=transaction_data.get('property_name', 'N/A'))
            ws.cell(row=row, column=4, value=transaction_data.get('location', 'N/A'))
            ws.cell(row=row, column=5, value=transaction_data.get('property_type', 'N/A'))
            ws.cell(row=row, column=6, value=transaction_data.get('transaction_type', 'N/A'))
            
            value = transaction_data.get('transaction_value', 0)
            ws.cell(row=row, column=7, value=value)
            if value > 0:
                ws.cell(row=row, column=7).number_format = '#,##0'
            
            # Check if it's a big deal
            from config import is_big_deal
            is_big = is_big_deal(transaction_data)
            ws.cell(row=row, column=8, value="Yes" if is_big else "No")
            if is_big:
                ws.cell(row=row, column=8).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            
            # Add URL
            ws.cell(row=row, column=9, value=transaction.get('url', 'N/A'))
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws)
    
    def _create_news_sheet(self, wb: Workbook, news: List[Dict]):
        """Create news sheet."""
        ws = wb.create_sheet("Market News")
        
        # Headers - start from row 1
        headers = ['Date', 'Source', 'Title', 'Summary', 'URL']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data - start from row 2
        for row, article in enumerate(news, 2):
            ws.cell(row=row, column=1, value=article['date'].strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=article['source'])
            ws.cell(row=row, column=3, value=article['title'])
            
            # Create summary from content
            content = article['content']
            summary = content[:200] + "..." if len(content) > 200 else content
            ws.cell(row=row, column=4, value=summary)
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
            
            # Add URL
            ws.cell(row=row, column=5, value=article.get('url', 'N/A'))
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws)
    
    def _create_statistics_sheet(self, wb: Workbook, report_data: Dict):
        """Create statistics sheet."""
        ws = wb.create_sheet("Statistics")
        
        # Title
        ws['A1'] = "Market Statistics"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        
        # Statistics
        stats = report_data['statistics']
        
        data = [
            ['Total Transactions', stats['total_transactions']],
            ['Total Transaction Value (HKD)', stats['total_transaction_value']],
            ['News Articles', stats['total_news_articles']],
            ['Sources Covered', len(stats['sources_covered'])],
            ['Sources', ', '.join(stats['sources_covered'])],
        ]
        
        for row, (label, value) in enumerate(data, 3):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            if isinstance(value, int) and value > 1000:
                ws.cell(row=row, column=2).number_format = '#,##0'
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws)
    
    def _create_big_deals_sheet(self, wb: Workbook, transactions: List[Dict]):
        """Create big deals sheet."""
        ws = wb.create_sheet("Big Deals")
        
        # Title
        ws['A1'] = "Big Deals Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:I1')
        
        # Baseline information
        ws['A3'] = "Big Deals Baseline:"
        ws['A3'].font = Font(bold=True)
        
        row = 4
        for prop_type, baseline in BIG_DEALS_BASELINE.items():
            ws.cell(row=row, column=1, value=f"{prop_type.title()}:")
            ws.cell(row=row, column=2, value=f"{baseline['min_value']:,} HKD")
            ws.cell(row=row, column=3, value=baseline['description'])
            row += 1
        
        # Headers for big deals data
        row += 2
        headers = [
            'Date', 'Source', 'Property Name', 'Location', 'Asset Type',
            'Transaction Type', 'Value (HKD)', 'Baseline Met', 'URL'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        # Big deals data
        from config import is_big_deal
        big_deals = [t for t in transactions if is_big_deal(t.get('transaction_data', {}))]
        
        for i, transaction in enumerate(big_deals, row + 1):
            transaction_data = transaction.get('transaction_data', {})
            
            ws.cell(row=i, column=1, value=transaction['date'].strftime('%Y-%m-%d'))
            ws.cell(row=i, column=2, value=transaction['source'])
            ws.cell(row=i, column=3, value=transaction_data.get('property_name', 'N/A'))
            ws.cell(row=i, column=4, value=transaction_data.get('location', 'N/A'))
            ws.cell(row=i, column=5, value=transaction_data.get('property_type', 'N/A'))
            ws.cell(row=i, column=6, value=transaction_data.get('transaction_type', 'N/A'))
            
            value = transaction_data.get('transaction_value', 0)
            ws.cell(row=i, column=7, value=value)
            if value > 0:
                ws.cell(row=i, column=7).number_format = '#,##0'
            
            # Check baseline
            prop_type = transaction_data.get('property_type', 'commercial')
            baseline = BIG_DEALS_BASELINE.get(prop_type, BIG_DEALS_BASELINE['commercial'])
            baseline_met = value >= baseline['min_value']
            ws.cell(row=i, column=8, value="Yes" if baseline_met else "No")
            
            # Add URL
            ws.cell(row=i, column=9, value=transaction.get('url', 'N/A'))
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws) 